from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl import Workbook
import qrcode
from io import BytesIO
import base64
import json
from datetime import datetime
import os
import urllib.parse

app = Flask(__name__)

# Use your secret key
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', '5f4dcc3b5aa765d61d8327deb882cf992b95990a9151374abd3e4357b8fa88b2')

# Database configuration - FORCE PostgreSQL for Vercel
def get_database_url():
    """Get database URL - IMPORTANT: Vercel MUST use PostgreSQL"""
    # First, check if we're running on Vercel
    is_vercel = os.environ.get('VERCEL', '0') == '1'
    
    # Get database URL from environment
    database_url = os.environ.get('DATABASE_URL')
    
    # If on Vercel but no DATABASE_URL, raise error
    if is_vercel and not database_url:
        raise ValueError(
            "DATABASE_URL environment variable is required for Vercel deployment. "
            "Please set it in your Vercel project settings."
        )
    
    # If we have a DATABASE_URL, ensure it's PostgreSQL
    if database_url:
        # Convert postgres:// to postgresql://
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        
        # Verify it's PostgreSQL
        if not database_url.startswith('postgresql://'):
            raise ValueError(f"Invalid database URL for Vercel. Must be PostgreSQL, got: {database_url[:50]}...")
        
        print("✅ Using PostgreSQL database (Neon)")
        return database_url
    
    # For local development ONLY - use SQLite
    print("⚠️ WARNING: Using SQLite for local development only")
    print("⚠️ This will NOT work on Vercel - set DATABASE_URL environment variable")
    
    # Create instance folder for SQLite
    instance_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
    os.makedirs(instance_path, exist_ok=True)
    
    # Set permissions for the folder
    try:
        os.chmod(instance_path, 0o777)
    except:
        pass
    
    db_path = os.path.join(instance_path, 'database.db')
    return f'sqlite:///{db_path}'

# Set database URL
app.config['SQLALCHEMY_DATABASE_URI'] = get_database_url()
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configure SQLAlchemy based on database type
if 'postgresql' in app.config['SQLALCHEMY_DATABASE_URI']:
    # PostgreSQL/Neon configuration
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_recycle': 300,
        'pool_pre_ping': True,
        'pool_size': 10,
        'max_overflow': 20,
        'connect_args': {
            'sslmode': 'require',
            'connect_timeout': 10,
            'application_name': 'robotica-app'
        }
    }
    print("✅ Configured for PostgreSQL with connection pooling")
else:
    # SQLite configuration
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_recycle': 300,
        'pool_pre_ping': True,
        'connect_args': {
            'check_same_thread': False,
            'timeout': 30
        }
    }
    print("⚠️ Configured for SQLite (local development only)")

db = SQLAlchemy(app)

# Models
class Admin(db.Model):
    __tablename__ = 'admin'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)

class Staff(db.Model):
    __tablename__ = 'staff'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    category = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Student(db.Model):
    __tablename__ = 'student'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(50), unique=True, nullable=False, index=True)
    name = db.Column(db.String(100), nullable=False)
    college = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    category = db.Column(db.String(50))
    qr_code = db.Column(db.Text)

class Evaluation(db.Model):
    __tablename__ = 'evaluation'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(50), nullable=False, index=True)
    staff_id = db.Column(db.Integer, nullable=False, index=True)
    category = db.Column(db.String(50), nullable=False)
    score = db.Column(db.Float, nullable=False)
    max_score = db.Column(db.Float, nullable=False)
    criteria_scores = db.Column(db.Text)
    comments = db.Column(db.Text)
    evaluated_at = db.Column(db.DateTime, default=datetime.utcnow)

# Helper function for Excel validation
def validate_excel_data(row):
    """Validate and clean Excel row data"""
    if len(row) < 2:
        return None
    
    student_id = str(row[0]).strip() if row[0] else None
    name = str(row[1]).strip() if len(row) > 1 and row[1] else None
    
    if not student_id or not name:
        return None
    
    if student_id.lower() in ['none', 'null', 'nan', '', 'id', 'student id']:
        return None
    
    college = str(row[2]).strip() if len(row) > 2 and row[2] else ""
    phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""
    
    category = None
    if len(row) > 4 and row[4]:
        cat_str = str(row[4]).strip().lower().replace(' ', '_')
        if cat_str in ['robo_race', 'robo_sumo', 'working_model']:
            category = cat_str
        elif 'race' in cat_str:
            category = 'robo_race'
        elif 'sumo' in cat_str:
            category = 'robo_sumo'
        elif 'working' in cat_str or 'model' in cat_str:
            category = 'working_model'
    
    return {
        'student_id': student_id,
        'name': name,
        'college': college,
        'phone': phone,
        'category': category
    }

# Create tables and admin user
def init_database():
    with app.app_context():
        try:
            print(f"Initializing database...")
            
            # Create all tables
            db.create_all()
            print("✅ Tables created successfully")
            
            # Test database connection
            db.session.execute('SELECT 1')
            print("✅ Database connection test successful")
            
            # Create default admin if not exists
            admin_count = Admin.query.count()
            if admin_count == 0:
                admin = Admin(
                    email="admin@robotica.com",
                    password=generate_password_hash("admin123")
                )
                db.session.add(admin)
                db.session.commit()
                print("✅ Default admin user created!")
            else:
                print(f"✅ Admin user already exists ({admin_count} admins)")
                
        except Exception as e:
            print(f"❌ Database initialization error: {e}")
            import traceback
            traceback.print_exc()
            
            # Rollback any failed transaction
            try:
                db.session.rollback()
            except:
                pass

# Initialize database
init_database()

# Routes
@app.route('/')
def home():
    if 'user_type' in session:
        if session['user_type'] == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('staff_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        try:
            # Check admin
            admin = Admin.query.filter_by(email=email).first()
            if admin and check_password_hash(admin.password, password):
                session['user_id'] = admin.id
                session['user_type'] = 'admin'
                session['email'] = admin.email
                flash('Login successful!', 'success')
                return redirect(url_for('admin_dashboard'))
            
            # Check staff
            staff = Staff.query.filter_by(email=email).first()
            if staff and check_password_hash(staff.password, password):
                session['user_id'] = staff.id
                session['user_type'] = 'staff'
                session['email'] = staff.email
                session['category'] = staff.category
                flash('Login successful!', 'success')
                return redirect(url_for('staff_dashboard'))
            
            flash('Invalid credentials!', 'danger')
        except Exception as e:
            flash(f'Login error: {str(e)}', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully!', 'info')
    return redirect(url_for('login'))

# Admin Routes
@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        total_students = Student.query.count()
        total_staff = Staff.query.count()
        total_evaluations = Evaluation.query.count()
        
        categories = ['robo_race', 'robo_sumo', 'working_model']
        category_counts = {}
        for cat in categories:
            category_counts[cat] = {
                'students': Student.query.filter_by(category=cat).count(),
                'staff': Staff.query.filter_by(category=cat).count(),
                'evaluations': Evaluation.query.filter_by(category=cat).count()
            }
    except Exception as e:
        flash(f'Error loading statistics: {str(e)}', 'danger')
        return render_template('admin_dashboard.html', 
                             total_students=0,
                             total_staff=0,
                             total_evaluations=0,
                             category_counts={})
    
    return render_template('admin_dashboard.html', 
                         total_students=total_students,
                         total_staff=total_staff,
                         total_evaluations=total_evaluations,
                         category_counts=category_counts)

@app.route('/admin/staff', methods=['GET', 'POST'])
def admin_staff():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            name = request.form['name']
            email = request.form['email']
            password = request.form['password']
            category = request.form['category']
            
            if Staff.query.filter_by(email=email).first():
                flash('Staff email already exists!', 'danger')
            else:
                staff = Staff(
                    name=name,
                    email=email,
                    password=generate_password_hash(password),
                    category=category
                )
                db.session.add(staff)
                db.session.commit()
                flash('Staff created successfully!', 'success')
        except Exception as e:
            flash(f'Error creating staff: {str(e)}', 'danger')
    
    try:
        staff_list = Staff.query.all()
    except:
        staff_list = []
    
    return render_template('admin_staff.html', staff_list=staff_list)

@app.route('/admin/delete_staff/<int:id>')
def delete_staff(id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        staff = Staff.query.get_or_404(id)
        db.session.delete(staff)
        db.session.commit()
        flash('Staff deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting staff: {str(e)}', 'danger')
    
    return redirect(url_for('admin_staff'))

@app.route('/admin/students', methods=['GET', 'POST'])
def admin_students():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        if 'excel_file' in request.files:
            file = request.files['excel_file']
            if file.filename.endswith(('.xlsx', '.xls')):
                try:
                    workbook = openpyxl.load_workbook(file, data_only=True)
                    sheet = workbook.active
                    
                    added = 0
                    skipped = 0
                    duplicates = 0
                    
                    # Start a fresh session for import
                    from sqlalchemy.orm import Session
                    import_session = Session(db.engine)
                    
                    try:
                        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                            try:
                                # Skip empty rows
                                if not any(row):
                                    skipped += 1
                                    continue
                                
                                # Validate data
                                data = validate_excel_data(row)
                                if not data:
                                    skipped += 1
                                    continue
                                
                                # Check if student already exists
                                existing = import_session.query(Student).filter_by(student_id=data['student_id']).first()
                                if existing:
                                    duplicates += 1
                                    continue
                                
                                # Generate QR code
                                qr = qrcode.QRCode(
                                    version=1,
                                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                                    box_size=10,
                                    border=4,
                                )
                                qr.add_data(data['student_id'])
                                qr.make(fit=True)
                                
                                img = qr.make_image(fill_color="black", back_color="white")
                                buffered = BytesIO()
                                img.save(buffered, format="PNG")
                                qr_base64 = base64.b64encode(buffered.getvalue()).decode()
                                
                                # Create student record
                                student = Student(
                                    student_id=data['student_id'],
                                    name=data['name'],
                                    college=data['college'],
                                    phone=data['phone'],
                                    category=data['category'],
                                    qr_code=qr_base64
                                )
                                
                                import_session.add(student)
                                added += 1
                                
                                # Commit every 10 records to avoid large transactions
                                if added % 10 == 0:
                                    import_session.commit()
                                    print(f"Committed {added} records so far...")
                                    
                            except Exception as e:
                                print(f"Error processing row {idx}: {e}")
                                import_session.rollback()
                                skipped += 1
                                continue
                        
                        # Final commit
                        import_session.commit()
                        flash(f'✅ Successfully imported {added} students! {skipped} rows skipped. {duplicates} duplicates skipped.', 'success')
                        
                    finally:
                        import_session.close()
                    
                except Exception as e:
                    print(f"Import error details: {e}")
                    import traceback
                    traceback.print_exc()
                    flash(f'❌ Error importing file: {str(e)}', 'danger')
            else:
                flash('Please upload a valid Excel file (.xlsx or .xls)', 'warning')
    
    try:
        students = Student.query.order_by(Student.student_id).all()
    except Exception as e:
        students = []
        print(f"Error loading students: {e}")
    
    return render_template('admin_students.html', students=students)

@app.route('/admin/delete_student/<int:id>')
def delete_student(id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        student = Student.query.get_or_404(id)
        db.session.delete(student)
        db.session.commit()
        flash('Student deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting student: {str(e)}', 'danger')
    
    return redirect(url_for('admin_students'))

@app.route('/admin/reports')
def admin_reports():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    category = request.args.get('category', 'all')
    
    try:
        if category == 'all':
            evaluations = Evaluation.query.all()
        else:
            evaluations = Evaluation.query.filter_by(category=category).all()
        
        categories = ['robo_race', 'robo_sumo', 'working_model']
        stats = {}
        for cat in categories:
            cat_evals = Evaluation.query.filter_by(category=cat).all()
            if cat_evals:
                scores = [e.score for e in cat_evals]
                stats[cat] = {
                    'count': len(cat_evals),
                    'avg': sum(scores) / len(scores),
                    'max': max(scores),
                    'min': min(scores)
                }
            else:
                stats[cat] = {
                    'count': 0,
                    'avg': 0,
                    'max': 0,
                    'min': 0
                }
    except Exception as e:
        evaluations = []
        stats = {}
        flash(f'Error loading reports: {str(e)}', 'danger')
    
    return render_template('admin_reports.html', 
                         evaluations=evaluations,
                         category=category,
                         stats=stats)

@app.route('/admin/export')
def export_reports():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    category = request.args.get('category', 'all')
    
    try:
        if category == 'all':
            evaluations = Evaluation.query.all()
        else:
            evaluations = Evaluation.query.filter_by(category=category).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Evaluations_{category}"
        
        headers = ['Student ID', 'Staff ID', 'Category', 'Score', 'Max Score', 'Comments', 'Evaluated At']
        ws.append(headers)
        
        for eval in evaluations:
            ws.append([
                eval.student_id,
                eval.staff_id,
                eval.category.replace('_', ' ').title(),
                eval.score,
                eval.max_score,
                eval.comments,
                eval.evaluated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"robotica_evaluations_{category}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buffer, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Error exporting reports: {str(e)}', 'danger')
        return redirect(url_for('admin_reports'))

# Staff Routes
@app.route('/staff/dashboard')
def staff_dashboard():
    if 'user_type' not in session or session['user_type'] != 'staff':
        return redirect(url_for('login'))
    
    try:
        staff_id = session['user_id']
        category = session['category']
        
        evaluations = Evaluation.query.filter_by(staff_id=staff_id).all()
        total_evaluated = len(evaluations)
        
        students_in_category = Student.query.filter_by(category=category).count()
    except Exception as e:
        evaluations = []
        total_evaluated = 0
        students_in_category = 0
        flash(f'Error loading dashboard: {str(e)}', 'danger')
    
    return render_template('staff_dashboard.html',
                         category=category,
                         total_evaluated=total_evaluated,
                         students_in_category=students_in_category,
                         evaluations=evaluations)

@app.route('/staff/evaluate', methods=['GET', 'POST'])
def staff_evaluate():
    if 'user_type' not in session or session['user_type'] != 'staff':
        return redirect(url_for('login'))
    
    staff_id = session['user_id']
    category = session['category']
    
    if request.method == 'POST':
        try:
            student_id = request.form['student_id']
            criteria_data = request.form.get('criteria_data', '{}')
            
            criteria = json.loads(criteria_data)
            total_score = sum(float(criteria[key]) for key in criteria)
            max_score = len(criteria) * 10
            
            student = Student.query.filter_by(student_id=student_id).first()
            if not student:
                flash('Student not found!', 'danger')
                return redirect(url_for('staff_evaluate'))
            
            if student.category != category:
                flash(f'Student belongs to {student.category.replace("_", " ").title()} category, not your assigned category!', 'danger')
                return redirect(url_for('staff_evaluate'))
            
            existing = Evaluation.query.filter_by(
                student_id=student_id,
                category=category
            ).first()
            
            if existing:
                flash('Student already evaluated! Re-evaluation is not allowed.', 'warning')
                return redirect(url_for('staff_evaluate'))
            
            evaluation = Evaluation(
                student_id=student_id,
                staff_id=staff_id,
                category=category,
                score=total_score,
                max_score=max_score,
                criteria_scores=criteria_data,
                comments=request.form.get('comments', '')
            )
            db.session.add(evaluation)
            db.session.commit()
            flash('Evaluation submitted successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error submitting evaluation: {str(e)}', 'danger')
        
        return redirect(url_for('staff_evaluate'))
    
    return render_template('staff_evaluate.html', category=category)

@app.route('/get_student/<student_id>')
def get_student(student_id):
    if 'user_type' not in session or session['user_type'] != 'staff':
        return jsonify({'exists': False})
    
    try:
        staff_category = session['category']
        student = Student.query.filter_by(student_id=student_id).first()
        
        if student:
            if student.category != staff_category:
                return jsonify({
                    'exists': True,
                    'wrong_category': True,
                    'student_category': student.category.replace('_', ' ').title(),
                    'staff_category': staff_category.replace('_', ' ').title()
                })
            
            evaluated = Evaluation.query.filter_by(
                student_id=student_id,
                category=staff_category
            ).first()
            
            return jsonify({
                'exists': True,
                'wrong_category': False,
                'name': student.name,
                'college': student.college,
                'phone': student.phone,
                'category': student.category,
                'already_evaluated': evaluated is not None
            })
    except Exception as e:
        return jsonify({'error': str(e), 'exists': False})
    
    return jsonify({'exists': False})

@app.route('/scanner')
def scanner():
    if 'user_type' not in session or session['user_type'] != 'staff':
        return redirect(url_for('login'))
    
    return render_template('scanner.html')

# Database test endpoint
@app.route('/test-db')
def test_db():
    try:
        db.session.execute('SELECT 1')
        
        admin_count = Admin.query.count()
        staff_count = Staff.query.count()
        student_count = Student.query.count()
        eval_count = Evaluation.query.count()
        
        return jsonify({
            'status': 'connected',
            'database_type': 'PostgreSQL' if 'postgresql' in app.config['SQLALCHEMY_DATABASE_URI'] else 'SQLite',
            'database_url_preview': app.config['SQLALCHEMY_DATABASE_URI'][:100] + '...',
            'counts': {
                'admins': admin_count,
                'staff': staff_count,
                'students': student_count,
                'evaluations': eval_count
            },
            'vercel_environment': os.environ.get('VERCEL', 'Not Vercel'),
            'has_database_url': bool(os.environ.get('DATABASE_URL'))
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'database_url': app.config['SQLALCHEMY_DATABASE_URI'][:100] + '...',
            'vercel_environment': os.environ.get('VERCEL', 'Not Vercel')
        })

# Health check
@app.route('/health')
def health():
    try:
        db.session.execute('SELECT 1')
        db_status = 'connected'
    except Exception as e:
        db_status = f'error: {str(e)}'
    
    return jsonify({
        'status': 'ok', 
        'timestamp': datetime.now().isoformat(),
        'database': db_status,
        'environment': 'production' if os.environ.get('VERCEL') else 'development'
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
